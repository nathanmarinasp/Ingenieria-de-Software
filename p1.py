#LEER ARCHIVOS CSV: 
#https://docs.python.org/es/3/library/csv.html
import csv 
import pandas 
df = pandas.read_csv('nombre_archivo.csv', sep = ',')

##################

with open('nombre_archivo.csv', newline = '') as csvfile:
    spamreader = csv.reader(csvfile, delimiter = '', quotechar = '|')
    for row in spamreader:
        print(', '.join(row))

#LEER ARCHIVOS EXCELL:
#https://programacion.net/articulo/como_trabajar_con_archivos_excel_utilizando_python_1419

import openpyxl
documento_excell = openpyxl.load_workbook('archivo_excell.xlsx')

#nombres de hojas 
nombre_archivo_excell.get_sheet_names()

#acceso a celdas
sheet = nombre_archivo_excell.get_sheet_by_name('sheet1')
print sheet['A2'].value

sheet.cell(row = 5, column = 2).value#mucho mas sencillo
print(sheet.cell(row = 5, column = 2))

#acceso a rango de celdas
multiples_celdas = sheet['A1':'B3']
for row in multiples_celdas:
    for cell in row:
        print cell.value

#acceso a todas las filas y columnas 
all_rows = sheet.rows
print all_rows[:]

all_columns = sheet.columns
print all_columns[:]

#REGRESIONES
from sklearn.linear_model import LinearRegression

df = pandas.read_csv('archivo_csv.csv')
x = df.x.values.reshape(-1,1)
y = df.y.values.reshape(-1,1)

reg = LinearRegression().fit(x, y)
reg.predict([[7]])

####################

import matplotlib.pyplot as plt
import numpy as np
from sklearn.linear_model import LinearRegression

#datos que queremos introducir (ejemplo)
mtros_cuadrados = np.array([[7,5,5,6,7,8,4,6,2,1,4,5]]).T
precios = np.array([[1800, 1600, 1500, 1000,5555,23555,1244]]).T

#regresor lineal
modelo = LinearRegression()
modelo.fit(mtros_cuadrados, precios) #datos que queremos meter en la grafica

#predecir salida
valores_a_predecir = np.array([[50,60,55,80,120,130,145,150]]).T
y_predichos = modelo.predict(valores_a_predecir)